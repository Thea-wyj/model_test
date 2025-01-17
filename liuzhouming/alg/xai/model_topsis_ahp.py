import numpy as np
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

class TOPSIS:
    def __init__(self, b, s, i, c):
        self.RI = (0, 0, 0.58, 0.9, 1.12, 1.24, 1.32, 1.41, 1.45, 1.49)
        self.b = b
        self.s = s
        self.i = i
        self.c = c

    def norm_data(self, x):  # 规范化
        sum = np.sum(x, axis=0)
        sum[np.where(sum==0)] = 1
        return x / sum

    # def zscore(self, x):
    #     return (x - x.mean(axis=0)) / x.std(axis=0)

    def ideal_solution(self, x, s):  # 计算理想解和负理想解,s中-1表示成本型指标越小越好，1表示收益型指标越大越好
        xt = x * s
        # 理想解
        vp = np.amax(xt, 0) * s
        vn = np.amin(xt, 0) * s
        # 计算距离
        sp = np.power((x-vp), 2)
        dp = np.power(sp.sum(axis=1), 1/2)
        sn = np.power((x-vn), 2)
        dn = np.power(sn.sum(axis=1), 1/2)
        # 计算接近度
        c = dn / (dp + dn)
        c = np.nan_to_num(c)

        data = {'D+':dp, 'D-':dn, 'C':c}
        df = pd.DataFrame(data)
        return df
        
    def printit(self, df_list, CR, eigen, df, W):
        wb = Workbook()
        
        sheet = wb.active
        sheet.title = '二级指标'
        row_count = 2
        for i in range(len(df_list)):
            sheet.append(['指标{}'.format(i+1)])
            sheet.append(['权重'])
            sheet.append(W[i].tolist())
            self.batch_format(sheet, 4, row_count, row_count+1, '00CED1')
            for r in dataframe_to_rows(df_list[i], index=True, header=True):
                sheet.append(r)
            self.batch_format(sheet, 4, row_count+2, row_count+8, 'FFE4B5')
            row_count = row_count + 10
        sheet = wb.create_sheet('一级指标')
        sheet.append(['一级指标权重'])
        sheet.append(eigen.tolist())
        sheet.append(['CR:', '{:<5f}'.format(CR)])
        self.batch_format(sheet, 4, 1, 3, 'FFE4B5')
        sheet.append(['一级指标接近度'])
        for r in dataframe_to_rows(df, index=True, header=True):
            sheet.append(r)
        self.batch_format(sheet, 4, 5, 11, '00CED1')

        wb.save('topsis_ahp.xls')
    
    def batch_format(self, ws, col, row1, row2, color):
        fille = PatternFill('solid', fgColor=color)  # 设置填充颜色
        for i in range(col):
            for j in range(row2-row1+1):
               cell = ws.cell(row=row1+j, column=i+1)
               cell.fill = fille

    def cal_weights(self, input_matrix):
        input_matrix = np.array(input_matrix)
        n, n1 = input_matrix.shape
        assert n == n1, '不是一个方阵'
        for i in range(n):
            for j in range(n):
                if np.abs(input_matrix[i, j] * input_matrix[j, i] - 1) > 1e-7:
                    raise ValueError('不是反互对称矩阵')

        eigenvalues, eigenvectors = np.linalg.eig(input_matrix)

        max_idx = np.argmax(eigenvalues)
        max_eigen = eigenvalues[max_idx].real
        eigen = eigenvectors[:, max_idx].real
        eigen = eigen / eigen.sum()

        if n > 9:
            CR = None
            warnings.warn('无法判断一致性')
        elif n == 1:
            CR = 0
        else:
            CI = (max_eigen - n) / (n - 1)
            CR = CI / self.RI[n - 1]
            if CR > 0.1:
                raise Exception("CR>0.1，一致性检验不通过")
        return max_eigen, CR, eigen

    def run(self):
        # 二级指标
        df_list = []
        W = []
        for i, bi in enumerate(self.b):
            y = self.norm_data(bi)  # 规范化
            max_eigen, CR, w = self.cal_weights(self.c[i])  # 权重
            v = w * y  # 加权规范化矩阵
            df_list.append(self.ideal_solution(v, self.s[i]))  # 计算理想距离差和接近度
            W.append(w)

        # 一级指标
        max_eigen, CR, weight = self.cal_weights(self.i)  # AHP计算一级指标的权重
        c_list = []
        for i, df in enumerate(df_list):
            c_list.append(df['C'].to_list())
        x = np.array(c_list).T
        # x = self.zscore(x)
        v = weight * x
        df = self.ideal_solution(v, np.ones(weight.shape[0]))
        
        self.printit(df_list, CR, weight, df, W)
        return df,W,weight
"""
指标顺序
一致性：auc
复杂性：规则数量，最大规则长度，规则总长度
明确性：覆盖率，类覆盖率，重合率，矛盾率
稳定性：预测稳定性，规则稳定性
"""
if __name__ == '__main__':
    # 二级指标矩阵 顺序mdrl brs sbrl
    b1 = np.array([[0.96], [0.948], [0.919]])
    b2 = np.array([[2.528, 7],
                   [3.093, 17],
                   [4.347, 31]])
    b3 = np.array([[0, 0],
                  [0.3529, 0.0645],
                  [0, 0.1508]])
    # b4 = np.array([[1, 1],
    #               [1, 1],
    #               [1, 1],
    #               [0, 0],
    #               [1, 1]])
    b = [b1, b2, b3]
    # 指标类型，1为收益型，-1为损失型
    s1 = np.array([1])
    s2 = np.array([-1, -1])
    s3 = np.array([-1, -1])
    # s4 = np.array([1, 1])
    s = [s1, s2, s3]
   # AHP 评价矩阵    
    # 准则层重要性矩阵
    criteria = np.array([[1, 3, 4],
                        [1/3, 1, 2],
                        [1/4, 1/2, 1]])
    # 方案层
    c1 = np.array([[1]])
    c2 = np.array([[1, 1],
                   [1, 1]])
    c3 = np.array([[1, 1],
                  [1, 1]])
    # c4 = np.array([[1, 3],
                #    [1/3, 1]])
    c = [c1, c2, c3]

    topsis = TOPSIS(b, s, criteria, c)
    df,W,weight=topsis.run()