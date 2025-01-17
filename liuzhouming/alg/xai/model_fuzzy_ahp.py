import numpy as np
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class AHP:
    def __init__(self, criteria, b, t, s, i, scores, filename):
        self.RI = (0, 0, 0.58, 0.9, 1.12, 1.24, 1.32, 1.41, 1.45, 1.49)
        self.criteria = criteria
        self.b = b  # AHP 评价矩阵
        self.t = t
        self.s = s
        self.i = i
        self.scores = scores  # 评分
        self.num_criteria = criteria.shape[0]
        self.num_project = b[0].shape[0]
        self.num_scores = scores.shape[0]
        self.filename = filename

    def AHP_weights(self, input_matrix):
        input_matrix = np.array(input_matrix)
        n, n1 = input_matrix.shape
        if n == 1:
            return 1, 0, np.array([1])
        assert n == n1, '不是一个方阵'
        for i in range(n):
            for j in range(n):
                if np.abs(input_matrix[i, j] * input_matrix[j, i] - 1) > 1e-7:
                    raise ValueError('不是反互对称矩阵')

        eigenvalues, eigenvectors = np.linalg.eig(input_matrix)

        max_idx = np.argmax(eigenvalues)
        max_eigen = eigenvalues[max_idx].real
        eigen = eigenvectors[:, max_idx].real
        eigen = eigen / eigen.sum()

        if n > 9:
            CR = None
            warnings.warn('无法判断一致性')
        elif n == 1:
            CR = 0
        else:
            CI = (max_eigen - n) / (n - 1)
            CR = CI / self.RI[n - 1]
            if CR > 0.1:
                raise Exception("CR>0.1，一致性检验不通过")
        return max_eigen, CR, eigen

    def run(self):
        # 准则层
        max_eigen, CR, criteria_eigen = self.AHP_weights(self.criteria)
        # 方案层
        max_eigen_list, CR_list, eigen_list = [], [], []
        for i, bi in enumerate(self.b):
            max_eigen, CR, eigen = self.AHP_weights(bi)
            max_eigen_list.append(max_eigen)
            CR_list.append(CR)
            eigen_list.append(eigen)

        # 模糊评价法
        self.cal_memberVector()
        R = np.empty(shape=(0, self.num_scores))  # 准则层的模糊评价矩阵
        for i in range(self.num_criteria):
            fi = eigen_list[i].dot(self.r[i]).reshape(1, -1)
            R = np.concatenate((R, fi), axis=0)
        S = R.dot(self.scores.T)  # 各个一级指标的得分
        F = criteria_eigen.dot(R)  # 一级指标模糊向量
        B = F.dot(self.scores.T)  # 总目标得分
        self.printout(R, S, F, B, criteria_eigen, max_eigen, CR, eigen_list, max_eigen_list, CR_list)
        return B, eigen_list, criteria_eigen, F

    def printout(self, R, S, F, B, criteria_eigen, max_eigen, CR, eigen_list, max_eigen_list, CR_list):
        workbook = Workbook()

        worksheet1 = workbook.active
        worksheet1.title = 'AHP权重'
        worksheet1.append(['一级指标权重'])
        worksheet1.append(criteria_eigen.tolist())
        worksheet1.append(['最大特征值:', '{:<5f}'.format(max_eigen), 'CR:', '{:<5f}'.format(CR)])
        self.batch_format(worksheet1, 4, 1, 3, '6495ED')
        worksheet1.append(['二级指标权重'])
        for i in range(len(eigen_list)):
            worksheet1.append(["准则{}子指标权重：".format(i + 1)])
            worksheet1.append(eigen_list[i].tolist())
            worksheet1.append(['最大特征值:', '{:<5f}'.format(max_eigen_list[i]), 'CR:', '{:<5f}'.format(CR_list[i])])
        self.batch_format(worksheet1, 4, 4, 13, 'FFA07A')

        worksheet1 = workbook.create_sheet('模糊评价矩阵')
        for i in range(len(self.r)):
            worksheet1.append(['b{}'.format(i + 1)])
            for a in self.r[i]:
                worksheet1.append(a)

        worksheet1 = workbook.create_sheet('模糊评价得分')
        worksheet1.append(['一级指标得分'])
        worksheet1.append(S.tolist())
        self.batch_format(worksheet1, 4, 1, 2, '6495ED')
        worksheet1.append(['一级指标模糊矩阵'])
        R = R.tolist()
        for i in range(len(R)):
            worksheet1.append(R[i])
        self.batch_format(worksheet1, 4, 3, 6, 'FFA07A')
        worksheet1.append(['总目标隶属度向量'])
        worksheet1.append(F.tolist())
        self.batch_format(worksheet1, 4, 7, 8, 'FFE4B5')
        worksheet1.append(['总目标得分'])
        worksheet1.append([B])
        self.batch_format(worksheet1, 4, 9, 10, '00CED1')

        # 保存
        workbook.save(self.filename + '.xls')

    def batch_format(self, ws, col, row1, row2, color):
        fille = PatternFill('solid', fgColor=color)  # 设置填充颜色
        for i in range(col):
            for j in range(row2 - row1 + 1):
                cell = ws.cell(row=row1 + j, column=i + 1)
                cell.fill = fille

    # f1隶属度函数
    def func_1(self, a, b, x):
        if x < a:
            return 0.5 * (1 + (x - a) / (x - b))
        elif x >= a and x < b:
            return 0.5 * (1 - (a - x) / (a - b))
        else:
            return 0

    # f2隶属度函数
    def func_2(self, a, b, c, x):
        if x < a:
            return 0.5 * (1 - (x - a) / (x - b))
        elif x >= a and x < b:
            return 0.5 * (1 + (a - x) / (a - b))
        elif x >= b and x < c:
            return (c - x) / (c - b)
        else:
            return 0

    # f3隶属度函数
    def func_3(self, a, b, c, x):
        if x >= a and x < b:
            return (x - a) / (b - a)
        elif x >= b and x < c:
            return 0.5 * (1 + (x - c) / (b - c))
        elif x >= c:
            return 0.5 * (1 - (c - x) / (b - x))
        else:
            return 0

    # f4隶属度函数
    def func_4(self, a, b, x):
        if x >= a and x < b:
            return 0.5 * (1 - (x - b) / (a - b))
        elif x >= b:
            return 0.5 * (1 + (b - x) / (a - x))
        else:
            return 0

    def loss_indicator_memdeg(self, score, u_now):
        '''
        计算隶属度向量,只能是四个等级的情况
        :param score: 专家打分
        :param u_now: 当前指标值
        :return: 隶属度向量
        '''
        mem_vector = []
        a = score[0]
        b = score[1]
        c = score[2]
        d = score[3]
        mem_vector.append(self.func_1(a, b, u_now))  # f1
        mem_vector.append(self.func_2(a, b, c, u_now))  # f2
        mem_vector.append(self.func_3(b, c, d, u_now))  # f3
        mem_vector.append(self.func_4(c, d, u_now))  # f4
        return mem_vector

    def benefit_indicator_memdeg(self, score, u_now):
        '''
        计算隶属度向量
        :param score: 专家打分
        :param u_now: 当前指标值
        :return: 隶属度向量
        '''
        mem_vector = []
        score = score[::-1]
        mem_vector = self.loss_indicator_memdeg(score, u_now)
        mem_vector = mem_vector[::-1]
        return mem_vector

    def cal_memberVector(self):
        R = []
        for j in range(len(self.t)):
            r = []
            for q in range(len(self.s[j])):
                if self.t[j][q] == 0:
                    r.append(self.loss_indicator_memdeg(self.s[j][q], self.i[j][q]))
                else:
                    r.append(self.benefit_indicator_memdeg(self.s[j][q], self.i[j][q]))
            R.append(r)
        self.r = R


if __name__ == '__main__':
    # AHP 评价矩阵
    # 准则层重要性矩阵
    criteria = np.array([[1, 3, 4],
                         [1 / 3, 1, 2],
                         [1 / 4, 1 / 2, 1]])

    # 方案层
    b1 = np.array([[1]])
    b2 = np.array([[1, 1],
                   [1, 1]])
    b3 = np.array([[1, 1],
                   [1, 1]])
    b = [b1, b2, b3]

    # 专家打分
    t1 = [1]  # t for 指标类型，1越大越好，0越小越好
    s1 = [[1, 0.7, 0.5, 0.3]]

    t2 = [0, 0]
    s2 = [[1, 2, 4, 6],
          [6, 12, 24, 30]]

    t3 = [0, 0]
    s3 = [[0.1, 0.2, 0.5, 0.7],
          [0.1, 0.2, 0.5, 0.7]]

    t = [t1, t2, t3]
    s = [s1, s2, s3]

    #
    # tree1
    i1 = [0.96]
    i2 = [2.528, 7]
    i3 = [0, 0]
    i = [i1, i2, i3]

    # tree2
    j1 = [0.948]
    j2 = [3.093, 17]
    j3 = [0.3529, 0.0645]
    j = [j1, j2, j3]
    # tree3
    k1 = [0.919]
    k2 = [4.347, 31]
    k3 = [0, 0.1508]
    k = [k1, k2, k3]

    # 模糊评价等级
    scores = np.array([90, 80, 65, 30])

    a1 = AHP(criteria, b, t, s, i, scores, "tree1").run()
    a2 = AHP(criteria, b, t, s, j, scores, "tree2l").run()
    a3 = AHP(criteria, b, t, s, k, scores, "tree3").run()
    print("tree1得分：{}".format(a1))
    print("tree2得分：{}".format(a2))
    print("tree3得分：{}".format(a3))
    # a are final weights
