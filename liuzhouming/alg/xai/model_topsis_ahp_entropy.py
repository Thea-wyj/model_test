import numpy as np
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


class TOPSIS_AHP_Entropy:
    def __init__(self, b, s, i):
        self.RI = (0, 0, 0.58, 0.9, 1.12, 1.24, 1.32, 1.41, 1.45, 1.49)
        self.b = b
        self.s = s
        self.i = i

    def norm_data(self, x):  # 规范化
        return x / np.sqrt(np.sum(np.power(x, 2), axis=0))

    def entropy_weight(self, x):  # 计算熵权
        if x.shape[1] == 1: return np.array([1.0])
        m = x.shape[0]
        k = 1 / np.log(m)
        s = np.nan_to_num(x * np.log(x))  # 将nan空值转换为0
        ej = -k * (s.sum(axis=0))  # 计算信息熵
        wi = (1 - ej) / np.sum(1 - ej)  # 计算每种指标的权重
        return wi

    def ideal_solution(self, x, s):  # 计算理想解和负理想解,s中-1表示成本型指标越小越好，1表示收益型指标越大越好
        xt = x * s
        # 理想解
        vp = np.amax(xt, 0) * s
        vn = np.amin(xt, 0) * s

        # 计算距离
        sp = np.nan_to_num(np.power((x - vp), 2))
        dp = np.power(sp.sum(axis=1), 1 / 2)
        sn = np.nan_to_num(np.power((x - vn), 2))
        dn = np.power(sn.sum(axis=1), 1 / 2)

        # 计算接近度
        c = dn / (dp + dn)
        c = np.nan_to_num(c)

        data = {'D+': dp, 'D-': dn, 'C': c}
        df = pd.DataFrame(data)
        return df

    def printit(self, df_list, CR, eigen, df, W):
        wb = Workbook()

        sheet = wb.active
        sheet.title = '二级指标'
        row_count = 2
        for i in range(len(df_list)):
            sheet.append(['指标{}'.format(i + 1)])
            sheet.append(['权重'])
            sheet.append(W[i].tolist())
            self.batch_format(sheet, 4, row_count, row_count + 1, '00CED1')
            for r in dataframe_to_rows(df_list[i], index=True, header=True):
                sheet.append(r)
            self.batch_format(sheet, 4, row_count + 2, row_count + 6, 'FFE4B5')
            row_count = row_count + 8
        sheet = wb.create_sheet('一级指标')
        sheet.append(['一级指标权重'])
        sheet.append(eigen.tolist())
        sheet.append(['CR:', '{:<5f}'.format(CR)])
        self.batch_format(sheet, 4, 1, 3, 'FFE4B5')
        sheet.append(['一级指标接近度'])
        for r in dataframe_to_rows(df, index=True, header=True):
            sheet.append(r)
        self.batch_format(sheet, 4, 5, 9, '00CED1')

        wb.save('topsis2.xls')

    def batch_format(self, ws, col, row1, row2, color):
        fille = PatternFill('solid', fgColor=color)  # 设置填充颜色
        for i in range(col):
            for j in range(row2 - row1 + 1):
                cell = ws.cell(row=row1 + j, column=i + 1)
                cell.fill = fille

    def cal_weights(self, input_matrix):
        input_matrix = np.array(input_matrix)
        n, n1 = input_matrix.shape
        assert n == n1, '不是一个方阵'
        for i in range(n):
            for j in range(n):
                if np.abs(input_matrix[i, j] * input_matrix[j, i] - 1) > 1e-7:
                    raise ValueError('不是反互对称矩阵')

        eigenvalues, eigenvectors = np.linalg.eig(input_matrix)

        max_idx = np.argmax(eigenvalues)
        max_eigen = eigenvalues[max_idx].real
        eigen = eigenvectors[:, max_idx].real
        eigen = eigen / eigen.sum()

        if n > 9:
            CR = None
            warnings.warn('无法判断一致性')
        elif n == 1:
            CR = 0
        else:
            CI = (max_eigen - n) / (n - 1)
            CR = CI / self.RI[n - 1]
            if CR > 0.1:
                raise Exception("CR>0.1，一致性检验不通过")
        return max_eigen, CR, eigen

    def run(self):
        # 二级指标
        df_list = []
        W = []
        for i, bi in enumerate(self.b):
            y = self.norm_data(bi)  # 规范化
            w = self.entropy_weight(y)  # 权重
            v = w * y  # 加权规范化矩阵
            df_list.append(self.ideal_solution(v, self.s[i]))  # 计算理想距离差和接近度
            W.append(w)

        # 一级指标
        max_eigen, CR, weight = self.cal_weights(self.i)  # AHP计算一级指标的权重
        c_list = []
        for i, df in enumerate(df_list):
            c_list.append(df['C'].to_list())
        x = np.array(c_list).T
        v = weight * x
        df = self.ideal_solution(v, np.ones(weight.shape[0]))

        self.printit(df_list, CR, weight, df, W)
        return df,W,weight

if __name__ == '__main__':
    # 二级指标矩阵 顺序mdrl brs sbrl
    # b1 = np.array([[0.96,2.528, 7,0.00001, 0.00001], [0.948,3.093, 17,0.3529, 0.0645], [0.919,4.347, 31,0.00001, 0.1508]])
    # b = [b1]
    # 指标类型，1为收益型，-1为损失型
    b1 = np.array([[0.96], [0.948], [0.919]])
    b2 = np.array([[2.528, 7],
                   [3.093, 17],
                   [4.347, 31]])
    b3 = np.array([[0, 0],
                   [0.3529, 0.0645],
                   [0, 0.1508]])
    b = [b1, b2, b3]
    # s1 = np.array([1,-1,-1,-1,-1])
    # s = [s1]
    s1 = np.array([1])
    s2 = np.array([-1, -1])
    s3 = np.array([-1, -1])
    s = [s1, s2, s3]
    # 一级指标的AHP评价矩阵
    i = np.array([[1, 3, 4],
                  [1 / 3, 1, 2],
                  [1 / 4, 1 / 2, 1]])

    topsis = TOPSIS_AHP_Entropy(b, s, i)
    topsis.run()
